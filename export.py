from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

pdfmetrics.registerFont(
    TTFont("Roboto", "assets/fonts/Roboto-Regular.ttf")
)
pdfmetrics.registerFont(
    TTFont("Roboto-Bold", "assets/fonts/Roboto-Bold.ttf")
)
pdfmetrics.registerFontFamily(
    "Roboto",
    normal="Roboto",
    bold="Roboto-Bold",
    italic="Roboto",
    boldItalic="Roboto-Bold",
)
styles = getSampleStyleSheet()

styles["Title"].fontName = "Roboto-Bold"
styles["Heading1"].fontName = "Roboto-Bold"
styles["Heading2"].fontName = "Roboto-Bold"
styles["Normal"].fontName = "Roboto"
def _money(value):
    return f"{value:,.2f} ₽".replace(",", " ")


# ======================================================
# PDF
# ======================================================

def build_pdf(cargo, calc, tariffs, rates, customs):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    styles["Title"].fontName = "Roboto-Bold"
    styles["Heading1"].fontName = "Roboto-Bold"
    styles["Heading2"].fontName = "Roboto-Bold"
    styles["Normal"].fontName = "Roboto"
    story = []

    story.append(Paragraph("Коммерческое предложение", styles["Title"]))
    story.append(
        Paragraph(
            datetime.now().strftime("Дата расчета: %d.%m.%Y %H:%M"),
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    story.append(Paragraph("Параметры груза", styles["Heading2"]))

    cargo_table = Table(
        [
            ["Товар", cargo["product_name"] or "Не указано"],
            ["Вес", f"{calc['total_weight']:.2f} кг"],
            ["Объем", f"{calc['volume']:.3f} м³"],
            ["Количество мест", cargo["qty"]],
            [
                "Размеры",
                f"{cargo['length']} × {cargo['width']} × {cargo['height']} мм",
            ],
            ["Инвойс", f"{cargo['invoice_usd']:,.2f} USD"],
        ],
        colWidths=[70 * mm, 90 * mm],
    )

    cargo_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Roboto-Bold"),   # первый столбец
                ("FONTNAME", (1, 0), (1, -1), "Roboto"),        # второй столбец

                ("FONTSIZE", (0, 0), (-1, -1), 10),

                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),

                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F3F3")),

                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    story.append(cargo_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Стоимость доставки", styles["Heading2"]))

    rows = [["Маршрут", "Стоимость", "Срок"]]

    for route, cost, days, _ in calc["results"]:
        rows.append(
            [
                route,
                _money(cost),
                f"{days} дн.",
            ]
        )

    tbl = Table(rows)

    tbl.setStyle(
        TableStyle(
            [
                # Шрифты
                ("FONTNAME", (0, 0), (-1, 0), "Roboto-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Roboto"),

                ("FONTSIZE", (0, 0), (-1, -1), 10),

                # Заголовок
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

                # Таблица
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ]
        )
    )

    story.append(tbl)

    if customs["enabled_flag"]:

        story.append(Spacer(1, 10))
        story.append(
            Paragraph("Таможенные платежи", styles["Heading2"])
        )

        customs_table = Table(
            [
                ["Таможенная стоимость", _money(calc["t_val"])],
                ["Пошлина", _money(calc["duty"])],
                ["НДС", _money(calc["vat"])],
                ["Таможенный сбор", _money(calc["fee"])],
                ["Итого", _money(calc["total_customs"])],
            ]
        )

        customs_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Roboto-Bold"),
                    ("FONTNAME", (1, 0), (1, -1), "Roboto"),

                    ("FONTSIZE", (0, 0), (-1, -1), 10),

                    ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),

                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F3F3")),

                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),

                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        story.append(customs_table)

    story.append(Spacer(1, 12))

    story.append(
        Paragraph(
            f"Полная себестоимость: {_money(calc['full_cost'])}",
            styles["Heading1"],
        )
    )

    doc.build(story)

    buffer.seek(0)

    return buffer


# ======================================================
# Excel
# ======================================================

def build_excel(cargo, calc, tariffs, rates, customs):

    wb = Workbook()

    ws = wb.active
    ws.title = "Расчет"

    bold = Font(bold=True)

    # -------------------------------------
    # Исходные данные
    # -------------------------------------

    ws["A1"] = "Исходные данные"
    ws["A1"].font = bold
    ws["A3"] = "Товар"
    ws["B3"] = cargo["product_name"]
    row = 3

    for key, value in cargo.items():
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = value
        row += 1

    row += 1

    ws.cell(row=row, column=1).value = "Курсы"
    ws.cell(row=row, column=1).font = bold
    row += 1

    for key, value in rates.items():
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = value
        row += 1

    row += 1

    ws.cell(row=row, column=1).value = "Тарифы"
    ws.cell(row=row, column=1).font = bold
    row += 1

    for key, value in tariffs.items():
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = value
        row += 1

    # -------------------------------------
    # Результаты
    # -------------------------------------

    result_sheet = wb.create_sheet("Результаты")

    result_sheet["A1"] = "Маршрут"
    result_sheet["B1"] = "Стоимость"
    result_sheet["C1"] = "Срок"
    result_sheet["D1"] = "Оплачиваемая база"

    for cell in result_sheet[1]:
        cell.font = bold

    r = 2

    for item in calc["results"]:
        for c, value in enumerate(item, start=1):
            result_sheet.cell(r, c).value = value
        r += 1

    r += 2

    result_sheet.cell(r, 1).value = "Полная себестоимость"
    result_sheet.cell(r, 2).value = calc["full_cost"]

    r += 2

    if customs["enabled_flag"]:
        result_sheet.cell(r, 1).value = "Таможенная стоимость"
        result_sheet.cell(r, 2).value = calc["t_val"]

        r += 1
        result_sheet.cell(r, 1).value = "Пошлина"
        result_sheet.cell(r, 2).value = calc["duty"]

        r += 1
        result_sheet.cell(r, 1).value = "НДС"
        result_sheet.cell(r, 2).value = calc["vat"]

        r += 1
        result_sheet.cell(r, 1).value = "Таможенный сбор"
        result_sheet.cell(r, 2).value = calc["fee"]

        r += 1
        result_sheet.cell(r, 1).value = "Итого таможня"
        result_sheet.cell(r, 2).value = calc["total_customs"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output