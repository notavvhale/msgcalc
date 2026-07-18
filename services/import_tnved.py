from openpyxl import load_workbook

from services.tnved_database import get_connection
from services.duty_parser import parse_duty


def import_tnved(path: str):
    wb = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    ws = wb["ТНВЭД"]

    conn = get_connection()

    # очищаем таблицу перед импортом
    conn.execute("DELETE FROM tnved")

    rows_count = 0
    errors = []

    for row_number, (code, description, duty_text, details) in enumerate(
        ws.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):

        if not code:
            continue

        try:

            duty = parse_duty(duty_text)

            conn.execute(
                """
                INSERT INTO tnved
                (
                    code,
                    description,

                    duty_text,

                    calculation_type,

                    percent_rate,

                    specific_rate,

                    specific_currency,

                    specific_unit,

                    specific_quantity,

                    details
                )
                VALUES
                (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
                """,
                (
                    str(code).strip(),

                    description.strip() if description else "",

                    duty["duty_text"],

                    int(duty["calculation_type"]),

                    duty["percent_rate"],

                    duty["specific_rate"],

                    duty["specific_currency"],

                    duty["specific_unit"],

                    duty["specific_quantity"],

                    details,
                ),
            )

            rows_count += 1

        except Exception as e:

            errors.append(
                {
                    "row": row_number,
                    "code": code,
                    "duty": duty_text,
                    "error": str(e),
                }
            )

    conn.commit()
    conn.close()

    print("=" * 60)
    print(f"Импортировано записей: {rows_count}")

    if errors:

        print(f"Ошибок: {len(errors)}")

        print("=" * 60)
        print("Первые ошибки:\n")

        for err in errors[:20]:

            print(
                f"Строка {err['row']}"
            )
            print(
                f"Код: {err['code']}"
            )
            print(
                f"Тариф: {err['duty']}"
            )
            print(
                f"Ошибка: {err['error']}"
            )
            print("-" * 60)

    else:

        print("Импорт завершён без ошибок.")